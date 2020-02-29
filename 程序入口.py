import sys,os
from PyQt5.QtWidgets import QApplication,QWidget,QTableWidgetItem,QMessageBox,QComboBox,QMenu,QFileDialog
from PyQt5.QtGui import QFont,QColor,QIcon,QCursor,QPixmap
from PyQt5.QtCore import Qt
import openpyxl
# 导入设计好的图形模块
import check
import ChildUi
import ChildUi2
import 核心逻辑 as hx

# 初始化数据库处理---------------------
import 标准库
bz=标准库.biaozhun_list
del bz[0]      #删除第一行标题
import 自定义库
diy=自定义库.diy_list
del diy[0]

data_list=bz+diy  #合并数据库
datacolums_list=hx.get_2colums(data_list)     #获取数据库前两列

# ---------------------------------------------

if __name__=='__main__':
    app=QApplication(sys.argv)
    mainwindow=check.Ui_Form()
    window=QWidget()
    mainwindow.setupUi(window)

    windowchild=ChildUi.Ui_Form()
    window2=QWidget()
    windowchild.setupUi(window2)

    windowchild2=ChildUi2.Ui_Dialog()
    window3=QWidget()
    windowchild2.setupUi(window3)


    def getRealPath(s):
        # 获取exe解压目录的绝对路径
        # import os, sys
        p = os.path.realpath(sys.path[0])
        p = p.replace(r'\base_library.zip', '')
        p = p + s
        return p

    def readToTableWidget(data,widget):
        # 向表格控件写入数据
        r=len(data)
        widget.setRowCount(r)    #设置表格行数
        for row_num in range(r):
            for col_num in range(7):            ###7列！！！！！
                text=data[row_num][col_num]
                if text=='None':                 #将None替换空值
                    text=''
                elif '00:00:00' in text:
                    text=text.split(' ')[0].replace('-','/')    #将时间格式化
                item = QTableWidgetItem(text)
                widget.setItem(row_num, col_num, item)  # 写入
        widget.resizeColumnsToContents()  # 自动列宽

    # 初始化[数据库表格控件]填入数据----------------------------
    tab1=windowchild.tableWidget    #获取子窗口表格控件
    tab2=windowchild.tableWidget_2
    readToTableWidget(bz,tab1)      #初始化
    readToTableWidget(diy,tab2)
    #------------------------------------------------------------

    def msg(self):
        reply = QMessageBox.warning(self,  # 使用infomation信息框
                                        "找不到数据库文件！",
                                        "请确保data.py文件在本目录下\n程序退出！",
                                        QMessageBox.Yes )
        sys.exit(1)

    def isnull_msg(self):
        reply = QMessageBox.warning(self,  # 使用infomation信息框
                                        "请输入查询内容",
                                        "请左侧输入要查询的规范编号或名称！",
                                        QMessageBox.Yes )

    def saveSucces_msg(self):
        reply = QMessageBox.information(self,  # 使用infomation信息框
                                        "消息提醒",
                                        "保存成功！",
                                        QMessageBox.Yes )



    def selectbox():
        #下拉列表控件改变，隐藏列功能
        select_id=mainwindow.comboBox.currentIndex()
        # print(select_id)
        if select_id==0:
            # 显示主要信息
            # print(select_id)
            # mainwindow.tableWidget.setColumnHidden(1, True)
            mainwindow.tableWidget.setColumnHidden(2, True)
            mainwindow.tableWidget.setColumnHidden(5, True)
            mainwindow.tableWidget.setColumnHidden(6, True)
            # mainwindow.tableWidget.resizeColumnsToContents()  # 自动列宽
        else:
            # mainwindow.tableWidget.setColumnHidden(1, False)
            mainwindow.tableWidget.setColumnHidden(2, False)
            mainwindow.tableWidget.setColumnHidden(5, False)
            mainwindow.tableWidget.setColumnHidden(6, False)
            # mainwindow.tableWidget.resizeColumnsToContents()  # 自动列宽

    def setfont():
        font = QFont('微软雅黑', 10)
        font.setBold(True)  # 设置字体加粗
        mainwindow.tableWidget.horizontalHeader().setFont(font)  # 设置表头字体
        # mainwindow.tableWidget.horizontalHeader().setLineWidth(1)
        mainwindow.tableWidget.horizontalHeader().setStyleSheet("QHeaderView::section { border-bottom: 1px solid gray; }")


    def clear_table():
        #清空表格
        mainwindow.tableWidget.setRowCount(0)
        # mainwindow.tableWidget.clearContents()



    def get_txtlist():
        # 获取文本 框内容
        l=[]
        t=mainwindow.plainTextEdit.toPlainText()  #将全部文字内容输出为一个字符串
        for i in t.split('\n'):
            # l.append(i)
            if i!="":
                l.append(i)
        # print(l)
        return l



    def gotorun():
        # 查询前清空表格内容
        clear_table()
        # 获取时间
        tiaojantime = mainwindow.dateEdit.date()
        txtlist=get_txtlist()
        if len(txtlist)<=0:         #如果空列表，则终止
            isnull_msg(window)
            print('如果空列表，终止')
            return '终止'
        res_list=[]
        print('这是文本框列表：',txtlist)
        for i in txtlist:           #遍历文本输入框列表开始
            i=i.replace(" ",'')
            if i=='':
                continue
            s=hx.strto(i)            #初始化字符串，去空格、字母大写、英文括号
            l=hx.get_rowlist(s,tiaojantime,datacolums_list,data_list)    #查询，返回结果列表。
            if hx.is_duowei(l):
                # print('多行结果')
                for r in l:
                    res_list.append(r)
            else:
                # print('一行结果')
                res_list.append(l)
            # print(res_list)

        # 输出到表格控件开始---------------------------------------------------
        count = 0
        for row in res_list:
            # print('aaaaaaaaaaaaaaaaaaaa',row)
            mainwindow.tableWidget.insertRow(count)
            mainwindow.tableWidget.setItem(count, 0, QTableWidgetItem(row[0]))  # 查询内容
            mainwindow.tableWidget.setItem(count, 1, QTableWidgetItem(row[1]))  # 名称检查
            mainwindow.tableWidget.setItem(count, 2, QTableWidgetItem(row[2]))  # 完整名称
            mainwindow.tableWidget.setItem(count, 3, QTableWidgetItem(row[3]))  # 废止状态
            mainwindow.tableWidget.setItem(count, 4, QTableWidgetItem(row[4]))  # 可用情况
            mainwindow.tableWidget.setItem(count, 5, QTableWidgetItem(row[5]))  # 生效时间
            mainwindow.tableWidget.setItem(count, 6, QTableWidgetItem(row[6]))  # 代替情况
            # mainwindow.tableWidget.resizeColumnsToContents()  #自动列宽

            # 格式化输出样式开始-------------------------------------------------------------------
            if row[1]!='正常':
                mainwindow.tableWidget.item(count, 0).setForeground(QColor(255, 0, 0))
                mainwindow.tableWidget.item(count, 1).setForeground(QColor(255, 0, 0))

            if row[1]=='编号错误 ':
                mainwindow.tableWidget.item(count, 0).setBackground(QColor(220, 220, 220))

            if row[7]=='自定义库':
                # mainwindow.tableWidget.item(count, 0).setBackground(QColor(0, 255, 0))
                info_path=getRealPath(r'\img\info.png')
                newitem=QTableWidgetItem(QIcon(info_path),row[0])
                newitem.setToolTip('自定义标准')
                mainwindow.tableWidget.setItem(count, 0, QTableWidgetItem(newitem))  # 查询内容
            if row[1] != '正常':
                mainwindow.tableWidget.item(count, 0).setForeground(QColor(255, 0, 0))


            if '相似值' in row[2]:
                # 模糊匹配到数值，插入combox，并建立信号
                mainwindow.tableWidget.setItem(count, 2, QTableWidgetItem(""))  # 完整名称设为空
                # print(row[2])
                rl=row[2].split('：')[1]
                rl=rl.replace('[','')
                rl=rl.replace(']','')
                rl=rl.replace("'",'')
                name_list=rl.split(',')
                # print(name_list)
                name_list.insert(0,row[0])
                # # 将列表装进combox
                d_comboxList['comb'+str(count)] = QComboBox()
                d_comboxList['comb' + str(count)].setToolTip('为你找了到多个值')
                d_comboxList['comb' + str(count)].addItems(name_list)
                mainwindow.tableWidget.setCellWidget(count, 0, d_comboxList['comb'+str(count)] )


                def change_row():
                    # 表格内下拉控件改变，更新当前行查询
                    row_index = mainwindow.tableWidget.currentRow()
                    tiaojantime = mainwindow.dateEdit.date()
                    s = d_comboxList['comb'+str(row_index)] .currentText()
                    # print('xxxxxxxxxxx选中下拉控件的值：',s)

                    l = hx.get_rowlist(s, tiaojantime, datacolums_list, data_list)  # 查询，返回结果列表。
                    # print('xxxxxxxxxxx查询结果列表：',l)
                    mainwindow.tableWidget.removeRow(row_index)
                    mainwindow.tableWidget.insertRow(row_index)

                    mainwindow.tableWidget.setItem(row_index, 0, QTableWidgetItem(l[0]))  # 查询内容
                    mainwindow.tableWidget.setItem(row_index, 1, QTableWidgetItem(l[1]))  # 名称检查
                    mainwindow.tableWidget.setItem(row_index, 2, QTableWidgetItem(l[2]))  # 完整名称
                    mainwindow.tableWidget.setItem(row_index, 3, QTableWidgetItem(l[3]))  # 废止状态
                    mainwindow.tableWidget.setItem(row_index, 4, QTableWidgetItem(l[4]))  # 可用情况
                    mainwindow.tableWidget.setItem(row_index, 5, QTableWidgetItem(l[5]))  # 生效时间
                    mainwindow.tableWidget.setItem(row_index, 6, QTableWidgetItem(l[6]))  # 代替情况

                d_comboxList['comb' + str(count)].currentTextChanged.connect(change_row)  #为每个combox创建信号

            count = count + 1
            mainwindow.tableWidget.resizeColumnsToContents()  #自动列宽


    # 创建一个全局字典，用来存储combox控件对象
    d_comboxList={}


    # 数据库检查
    if data_list is None:
        # print('没找到数据库，确保data.py文件在本exe目录下')
        msg(window)


    def saveNameMsg(self):
        fileName = QFileDialog.getSaveFileName(self,
                                               " 导出Excel文件",
                                               "./",
                                               "All Files (*);;Excel文件 Files (*.xlsx)")
        return fileName[0]



    # saveNameMsg(window)

    def get_tabletext(widget):
        # 获取表格内容到列表
        r=widget.rowCount()
        c=widget.columnCount()
        if r >= 1:
            # print(r,c)
            c_list=[]
            for i in range(r):
                r_list=[]
                for j in range(c):
                    # print(i, j)
                    cellText=widget.item(i, j).text()
                    r_list.append(cellText)
                c_list.append(r_list)
            # print(c_list)
            return c_list


    def DataToExcel():
        # 导出数据库到Excel
        date1 = get_tabletext(windowchild.tableWidget)
        date2 = get_tabletext(windowchild.tableWidget_2)
        # print('xxxxxxxxxxxxxxx',date1)
        wb = openpyxl.Workbook()
        del wb['Sheet']
        def createSheet(data,sheetname):
            sheet1 = wb.create_sheet(sheetname)
            title_range = ['标准编号', '标准名称', '实施时间', '废弃时间', '标准状态', '代替标准号', '备注']
            for i in range(7):
                sheet1.cell(1, i + 1).value = title_range[i]
            for i in range(len(data)):
                for j in range(7):
                    sheet1.cell(i + 2, j + 1).value = data[i][j]
        createSheet(date1,'标准库')
        createSheet(date2,'自定义库')

        fileName = saveNameMsg(window)
        if fileName=='':
            print('取消导出')
            return 0

        wb.save(fileName)
        saveSucces_msg(window)   #保存成功提醒


    def contextMenuEvent():
        if mainwindow.tableWidget.rowCount()!=0:  #只有表格有行时，才弹出右键菜单
            # 表格控件的右键菜单事件
            menu = QMenu()
            copyAction = menu.addAction("复制")
            outToExcelAction = menu.addAction("导出到Excel")
            p=QCursor.pos()                     #获取鼠标坐标
            action = menu.exec_(p)
            if action == copyAction:
                r = mainwindow.tableWidget.currentRow()  # 获取当前选中的行
                if r!=-1:    #如果表格里有行信息
                     t=mainwindow.tableWidget.currentItem().text()
                     clipboard=QApplication.clipboard()
                     # print('复制', t)
                     clipboard.setText(t)
            elif action==outToExcelAction:
                # print('导出Excel')
                res_list=get_tabletext(mainwindow.tableWidget)
                rows = len(res_list)       #获取结果行数

                wb = openpyxl.Workbook()
                del wb['Sheet']
                sheet1 = wb.create_sheet('查询结果')
                title_range=['查询内容','名称检查','完整名称','废止状态','可用情况','生效时间','替代情况']
                for i in range(7):
                    sheet1.cell(1, i+1).value = title_range[i]
                for i in range(rows):
                    for j in range(7):
                        sheet1.cell(i+2, j + 1).value = res_list[i][j]

                fileName=saveNameMsg(window)
                if fileName == '':
                    print('取消导出')
                    return 0
                wb.save(fileName)
                saveSucces_msg(window)  # 保存成功提醒

    mainwindow.tableWidget.setContextMenuPolicy(Qt.CustomContextMenu)               #激活自定义右键菜单策略
    mainwindow.tableWidget.customContextMenuRequested.connect(contextMenuEvent)     #设置信号触发


    setfont()    # 设置表头格式
    selectbox()  # 显示combox内容

    #动作信号-------------------------------------------------------
    windowchild2.pushButton.clicked.connect(window3.close)          #3.关于窗体-确定按钮
    windowchild.pushButton_4.clicked.connect(DataToExcel)           #2.数据库窗口-导出Excel按钮
    mainwindow.pushButton_2.clicked.connect(gotorun)                #1.查询
    mainwindow.pushButton.clicked.connect(clear_table)              #1.清除按钮
    mainwindow.comboBox.currentTextChanged.connect(selectbox)       #1.下拉列表控件
    mainwindow.dateEdit.dateChanged.connect(gotorun)                #1.日期控件
    # mainwindow.plainTextEdit.textChanged.connect(gotorun)           #1.文本框控件
    mainwindow.plainTextEdit.blockCountChanged.connect(gotorun)           #1.
    mainwindow.pushButton_3.clicked.connect(window2.show)           #1.数据库按钮
    mainwindow.pushButton_4.clicked.connect(window3.show)           #1.关于按钮
    #---------------------------------------------------------------
    # 按钮格式化
    butt_path = getRealPath(r'\img\button-alert.png')          #获取exe目录绝对路径
    butt_path=butt_path.replace('\\','/')                         #图片路径处理1
    style="QPushButton{border-image: url("+butt_path+")}"    #图片路径处理1
    mainwindow.pushButton_4.setStyleSheet(style)
    mainwindow.pushButton_4.setToolTip('关于')

    pix_path = getRealPath(r'\img\校对2.png')
    # print('xxxxxxxxxxxxx',pix_path)
    pix = QPixmap(pix_path)                    #关于窗体-logo图片
    windowchild2.label.setPixmap(pix)
    windowchild2.label.setScaledContents(True)


    window3.setWindowTitle('关于Build time: 2020.2.29')
    gy_path = getRealPath(r'\img\button-alert.png')
    window3.setWindowIcon(QIcon(gy_path))

    window2.setWindowTitle('数据库')

    window.setWindowTitle('华信文本设计依据校对工具2.0')
    app.setWindowIcon(QIcon(pix_path))
    window.show()
    sys.exit(app.exec_())

