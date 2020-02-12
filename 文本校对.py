import openpyxl
import datetime


# 读取excel文档,转入列表
def read_exceldata():
    try:
        wb = openpyxl.load_workbook('data.xlsx')
        sheet = wb['Sheet1']
        rows = sheet.max_row
        cols = sheet.max_column
        # 存放excel数据到列表
        list_data = []
        for row in sheet.rows:
            list_temp = []
            for cell in row:
                list_temp.append(cell.value)
            list_data.append(list_temp)
        # 去空格
        for row in list_data:
            if  row[0] or row[1] is not  None:
                row[0] = row[0].replace(" ", "")
                row[1] = row[1].replace(" ", "")
        return list_data
    except FileNotFoundError:
        # print('没找到数据库，确保data.xlsx文件在本exe目录下')
        return None


# 丢进编号或名称、数据列表，返回查到的行的列表
def query_list(t,data):
    for row in data:
        # print(row)
        if t in row:
            # print(row)
            return row

# 丢进有效行，返回计算好时间的结果列表
def  jieguo_list(r,time_tiaojian):
    if r is None:
        print("查无此条目，检查输入编号或名称是否正确")
    else:
        # print(r)
        bianhao = r[0]                            # 标准编号
        biaozhun_name = r[1]                      # 标准名称
        shishi_time = r[2]                        # 实施时间
        feizhi_time = r[3]                        # 废弃时间
        biaozhun_zhuangtai = r[4]                 # 标准状态
        daitiqingkuang = r[5]                     # 代替情况
        update_off = r[6]                         # 更新管控
        wanzheng_name = bianhao + biaozhun_name   # 完整名称

         # 有效时间连接处理
        if feizhi_time is None:
            feizhi_time2=""
        else:
            feizhi_time2=str(feizhi_time).split(" ")[0].replace("-",".")
        shishi_time2=str(shishi_time).split(" ")[0].replace("-",".")

        youxiao_time = shishi_time2 + "~" + feizhi_time2
        # 返回值列表设计：["完整名称","规范状态","可用情况","有效时间","更替情况","更新管控"]
        list_res = []
        if biaozhun_zhuangtai == "作废":
            # 废止
            if time_tiaojian < shishi_time:
                # a.情况,废止-不可用
                list_res = [wanzheng_name, "废止", "不可用，此时规范还未实施", youxiao_time, daitiqingkuang, update_off]
                return list_res
            elif time_tiaojian >= shishi_time and time_tiaojian < feizhi_time:
                # b.情况,废止-可用
                list_res = [wanzheng_name, "废止", "可用", youxiao_time, daitiqingkuang, update_off]
                return list_res
            elif time_tiaojian >= feizhi_time:
                # c.情况,废止-不可用
                list_res = [wanzheng_name, "废止", "不可用，规范已废弃", youxiao_time, daitiqingkuang, update_off]
                return list_res
        else:
            # 现行
            if time_tiaojian < shishi_time:
                # d.情况,有效-不可用
                list_res = [wanzheng_name, "有效", "不可用，此时规范还未实施", youxiao_time, daitiqingkuang, update_off]
                return list_res
            elif time_tiaojian >= shishi_time:
                # e.情况,有效-可用
                list_res = [wanzheng_name, "有效", "可用", youxiao_time, daitiqingkuang, update_off]
                return list_res

# 输入处理，尝试提取编号,提取失败则原输入去空格处理
def get_bianhao(t):
    try:
        import re
        t = t.replace(" ", "")
        # p = re.compile(r'[GgJjYy][AaBbGgDd][ /JjTt]?[Tt]? ?\d{2,6}\.?\d{0,4}-\d{4}')       #旧的备份！！！
        p = re.compile(r'[GgJjYy][AaBbGgDdFf][ /JjTt]?[Tt]? ?/?[Tt]? ?\d{2,6}\.?\d{0,4}-\d{2,4}')

        r = p.findall(t)
        return r[0]
    except IndexError:
        return t

def main_query(t,time_tiaojian,data):
    r = get_bianhao(t)
    r2 = query_list(r, data)
    if r2 is None:
        # print('规范不存在，请检查输入是否正确')
        return '规范不存在，或输入有误'
    else:
        # print('成功查到目标：', r2)
        r3 = jieguo_list(r2, time_tiaojian)
        # print('成功计算结果：', r3)
        return r3


# r=query_list(t)
# res=jieguo_list(r)
# print(res)

# ----------------------------------------以下是测试------------------------------------

# # 编号获得测试---------------------------------------------
# # ---------------------------------------------------------
# t="YD/T 5175-2009通信局（站）防雷与接地工程验收规范"
# r=get_bianhao(t)
# print(r)
# # -----------------------------------------------------------

# 查询规范所在的行，返回整行信息到列表---------------------------------------------
# ---------------------------------------------------------

# t = "YD/T 5175-2009通信局（站）防雷与接地工程验收规范"
# t_list=["YD/T 5175-2009通信局（站）防雷与接地工程验收规范","YD/T 5175-3009通信局（站）防雷与接地工程验收规范","YD/T 5175-2009通信局（站）防雷与接地工程验收规范","YD/T 5175-2009通信局（站）防雷与接地工程验收规范"]
# str_p = '2012-01-30'
# time_tiaojian = datetime.datetime.strptime(str_p, '%Y-%m-%d')
# data = read_exceldata()
# res_list=[]
# for t in t_list:
#     r=main_query(t,time_tiaojian)
#     res_list.append(r)
# print(res_list)

# for i in res_list:
#     if isinstance(i, str):
#         print(i)


# -----------------------------------------------------------

